import openpyxl
import re
import json
import os

def slugify(text):
    text = text.lower()
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return text.strip('-')

def get_cat_id(cat, prod_name):
    if cat:
        c = cat.strip().lower()
        if c == 'tarugo': return 'tarugos'
        if c == 'tirafondo': return 'tirafondos'
        return c
    
    # Heuristics based on product name
    p = prod_name.upper()
    if 'WALL' in p: return 'autoperforantes'
    if 'FIX' in p: return 'autoperforantes'
    if 'DRY' in p: return 'autoperforantes'
    if 'PAM' in p: return 'autoperforantes'
    if 'ALAS' in p: return 'autoperforantes'
    if 'HT' in p or 'HEX T' in p: return 'autoperforantes'
    if 'GANCHO' in p: return 'ganchos'
    if 'PITON' in p: return 'pitones'
    if 'CLAVO' in p: return 'clavos'
    if 'REMACHE' in p: return 'remaches'
    if 'TUERCA' in p or 'TCA' in p: return 'tuercas'
    if 'VARILLA' in p: return 'varillas'
    if 'ARANDELA' in p: return 'arandelas'
    if 'BULON' in p: return 'bulones'
    if 'TIR' in p: return 'tirafondos'
    return 'otros'

def run():
    xlsx_file = "Archivo Ejemplo.xlsx"
    if not os.path.exists(xlsx_file):
        print(f"Error: No se encontró el archivo '{xlsx_file}' en el directorio raíz.")
        return

    print(f"Cargando {xlsx_file}...")
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    if 'Old' not in wb.sheetnames:
        print("Error: No se encontró la hoja 'Old' en el archivo Excel.")
        return

    sheet = wb['Old']
    print(f"Procesando {sheet.max_row} filas de la hoja 'Old'...")

    products_dict = {}

    for r in range(2, sheet.max_row + 1):
        code = sheet.cell(row=r, column=3).value
        cat = sheet.cell(row=r, column=4).value
        prod = sheet.cell(row=r, column=5).value
        brand = sheet.cell(row=r, column=6).value
        medida = sheet.cell(row=r, column=7).value
        presentacion = sheet.cell(row=r, column=8).value
        desc_pag = sheet.cell(row=r, column=9).value
        caracteristicas = sheet.cell(row=r, column=10).value
        usos = sheet.cell(row=r, column=11).value
        descripcion = sheet.cell(row=r, column=12).value

        # Ignorar filas sin nombre de producto
        if not prod:
            continue

        prod_name = str(prod).strip()
        prod_id = slugify(prod_name)

        if prod_id not in products_dict:
            brand_str = str(brand).strip() if brand else ""
            brand_str = brand_str.replace("Tel®", "Telé").replace("Tel\xae", "Telé")
            products_dict[prod_id] = {
                "id": prod_id,
                "name": prod_name,
                "category": get_cat_id(cat, prod_name),
                "brand": brand_str,
                "description": "",
                "features": "",
                "applications": "",
                "variants_dict": {}
            }

        p = products_dict[prod_id]

        # Actualizar categoría si es que previamente se había resuelto como 'otros' o vacía
        resolved_cat = get_cat_id(cat, prod_name)
        if p["category"] == "otros" and resolved_cat != "otros":
            p["category"] = resolved_cat

        # Acumular descripciones y campos dinámicos
        if desc_pag and not p["description"]:
            p["description"] = str(desc_pag).strip()
        if caracteristicas and not p["features"]:
            p["features"] = str(caracteristicas).strip()
        if usos and not p["applications"]:
            p["applications"] = str(usos).strip()

        # Construir y registrar variante
        var_id = str(code).strip() if code else f"{prod_id}-var-{len(p['variants_dict'])+1}"
        medida_str = str(medida).strip() if medida else "Estándar"
        pres_str = str(presentacion).strip() if presentacion else ""
        desc_str = str(descripcion).strip() if descripcion else f"{prod_name} {medida_str}"

        # Evitar duplicados exactos dentro del mismo producto
        if var_id not in p["variants_dict"]:
            p["variants_dict"][var_id] = {
                "id": var_id,
                "medida": medida_str,
                "presentacion": pres_str,
                "descripcion": desc_str
            }
        else:
            # Rellenar campos faltantes en variantes existentes
            v_exist = p["variants_dict"][var_id]
            if not v_exist["medida"] and medida_str:
                v_exist["medida"] = medida_str
            if not v_exist["presentacion"] and pres_str:
                v_exist["presentacion"] = pres_str
            if (not v_exist["descripcion"] or v_exist["descripcion"] == f"{prod_name} {medida_str}") and desc_str:
                v_exist["descripcion"] = desc_str

    # Convertir dict a lista ordenada alfabéticamente
    sorted_prods = sorted(products_dict.values(), key=lambda x: x["name"])

    # Distribuir las 32 imágenes de catálogo cíclicamente
    # Las imágenes están numeradas de la siguiente manera:
    # 1: imagenes/Imagenes y tornillos para catalogo 900px.jpg
    # 2 a 32: imagenes/Imagenes y tornillos para catalogo 900px{2-32}.jpg
    print(f"Distribuyendo las 32 imágenes entre los {len(sorted_prods)} productos únicos...")
    for idx, p in enumerate(sorted_prods):
        img_num = (idx % 32) + 1
        if img_num == 1:
            p["image"] = "imagenes/Imagenes y tornillos para catalogo 900px.jpg"
        else:
            p["image"] = f"imagenes/Imagenes y tornillos para catalogo 900px{img_num}.jpg"

        # Aplicar fallbacks a campos vacíos
        if not p["description"]:
            p["description"] = f"{p['name']}. Consulte medidas y disponibilidad."
        if not p["features"]:
            p["features"] = "Consulte características técnicas de este producto."
        if not p["applications"]:
            p["applications"] = "Ideal para fijaciones generales y uso industrial según medida."

        # Convertir variants_dict a lista
        p["variants"] = list(p["variants_dict"].values())
        del p["variants_dict"]

    # Guardar productos.json
    output_path = "productos.json"
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(sorted_prods, f, indent=2, ensure_ascii=False)
        print(f"¡Éxito! Catálogo generado en {output_path} con {len(sorted_prods)} productos.")
    except Exception as e:
        print(f"Error al guardar el archivo JSON: {e}")

if __name__ == "__main__":
    run()
